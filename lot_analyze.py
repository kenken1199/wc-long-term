import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
from scipy import stats

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
except ImportError:
    import types
    _stub = object
    tk = types.SimpleNamespace(Tk=_stub, Toplevel=_stub, Frame=_stub, StringVar=_stub, IntVar=_stub)
    ttk = types.SimpleNamespace(Frame=_stub, Button=_stub, Label=_stub, Entry=_stub, Treeview=_stub, Scrollbar=_stub, Spinbox=_stub, Notebook=_stub)
    messagebox = None
    filedialog = None
import datetime
from io import BytesIO
import os

import font_utils

font_utils.setup_japanese_font()

# =========================
# ■ 定数
# =========================
MIN_OK_COUNT = 2  # 統計分析に必要な最小OKデータ数
_RANK_LABEL_MAP = {"2": "OK", "1": "軽量", "E": "過量", "0": "２個乗り"}

# =========================
# ■ ロットプレビューダイアログ
# =========================
class LotPreviewDialog(tk.Toplevel):

    def __init__(self, parent, df, hinshoku_num=None, product_name=""):
        super().__init__(parent)
        self.title("ロット分割プレビュー")
        self.result = None
        self.df = df.copy()
        self.hinshoku_num = hinshoku_num
        self.product_name = product_name
        self.resizable(True, True)
        self.grab_set()

        # --- しきい値入力 ---
        frame_top = ttk.Frame(self, padding=10)
        frame_top.pack(fill="x")

        ttk.Label(frame_top, text="分割しきい値（分）:").pack(side="left")
        self.threshold_var = tk.IntVar(value=30)
        ttk.Spinbox(
            frame_top, from_=1, to=480,
            textvariable=self.threshold_var, width=6,
        ).pack(side="left", padx=5)

        # プリセットボタン（クリックで即時反映）
        presets = [("15分", 15), ("30分", 30), ("1時間", 60), ("2時間", 120), ("4時間", 240)]
        frame_presets = ttk.Frame(frame_top)
        frame_presets.pack(side="left", padx=(10, 0))
        for label, minutes in presets:
            ttk.Button(
                frame_presets, text=label, width=6,
                command=lambda m=minutes: self._set_threshold(m),
            ).pack(side="left", padx=2)

        self.lot_label_var = tk.StringVar()
        ttk.Label(frame_top, textvariable=self.lot_label_var, foreground="navy").pack(side="left", padx=15)

        # スピンボックスへの直接入力もリアルタイムに反映
        self.threshold_var.trace_add("write", self._on_threshold_change)

        # --- Treeview ---
        frame_tree = ttk.Frame(self, padding=10)
        frame_tree.pack(fill="both", expand=True)

        first_col = "製品名" if product_name else "品種番号"
        cols = (first_col, "ロット", "開始時刻", "終了時刻", "総件数", "OKデータ件数", "状態")
        self.tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=10)
        col_widths = {first_col: 110, "ロット": 70, "開始時刻": 150, "終了時刻": 150,
                      "総件数": 80, "OKデータ件数": 110, "状態": 110}
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor="center", width=col_widths[col])

        # 警告色タグ（OKデータ不足のロットを赤系で強調）
        self.tree.tag_configure("skip", background="#FFE4E1", foreground="#9C0006")

        scrollbar = ttk.Scrollbar(frame_tree, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.tree.bind("<Double-1>", self._on_lot_double_click)

        ttk.Label(self, text="💡 行をダブルクリックするとロット詳細を確認できます",
                  foreground="gray", font=("", 9), padding=(10, 2)).pack(fill="x")

        # --- 注意書きラベル ---
        self.warn_label_var = tk.StringVar()
        warn_label = ttk.Label(self, textvariable=self.warn_label_var,
                               foreground="#9C0006", padding=(10, 0))
        warn_label.pack(fill="x")

        # --- ボタン ---
        frame_btn = ttk.Frame(self, padding=(10, 5, 10, 10))
        frame_btn.pack(fill="x")
        ttk.Button(frame_btn, text="この分割でOK", command=self._on_ok).pack(side="left", padx=5)
        ttk.Button(frame_btn, text="ロット分割しない", command=self._on_no_split).pack(side="left", padx=5)
        ttk.Button(frame_btn, text="キャンセル", command=self._on_cancel).pack(side="right", padx=5)

        self._update_preview()

        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")

    def _compute_lots(self, threshold_min):
        df = self.df.sort_values("日付時刻").copy()
        df["時間差(分)"] = df["日付時刻"].diff().dt.total_seconds() / 60
        df["ロット"] = (df["時間差(分)"] > threshold_min).cumsum() + 1
        return df

    def _set_threshold(self, minutes):
        self.threshold_var.set(minutes)

    def _on_threshold_change(self, *_args):
        try:
            self.threshold_var.get()
        except tk.TclError:
            return  # 入力途中の空欄・不正値は無視
        self._update_preview()

    def _update_preview(self):
        try:
            threshold = self.threshold_var.get()
        except tk.TclError:
            return
        df = self._compute_lots(threshold)

        for item in self.tree.get_children():
            self.tree.delete(item)

        lot_count = df["ロット"].nunique()
        skip_count = 0

        if self.product_name:
            hinshoku_display = self.product_name
        elif self.hinshoku_num is not None:
            hinshoku_display = str(self.hinshoku_num)
        else:
            hinshoku_display = "-"

        for lot, group in df.groupby("ロット"):
            start = group["日付時刻"].min()
            end = group["日付時刻"].max()
            total = len(group)
            ok_count = int((group["ランクコード"] == "2").sum())

            if ok_count < MIN_OK_COUNT:
                state = "⚠ スキップ予定"
                tags = ("skip",)
                skip_count += 1
            else:
                state = "✓ 分析対象"
                tags = ()

            self.tree.insert("", "end", tags=tags, values=(
                hinshoku_display,
                f"ロット{lot}",
                start.strftime("%Y-%m-%d %H:%M") if pd.notna(start) else "-",
                end.strftime("%Y-%m-%d %H:%M") if pd.notna(end) else "-",
                total,
                ok_count,
                state,
            ))

        # 上部ラベル
        if skip_count > 0:
            self.lot_label_var.set(
                f"→ {lot_count} ロット検出（うち {skip_count} ロットはOKデータ不足でスキップ予定）"
            )
            self.warn_label_var.set(
                f"※ OKデータが {MIN_OK_COUNT} 件未満のロットは統計計算ができないため、"
                f"Excelファイルは作成されません。しきい値を変更してロットを統合することも検討してください。"
            )
        else:
            self.lot_label_var.set(f"→ {lot_count} ロット検出")
            self.warn_label_var.set("")

        self._df_with_lots = df

    def _on_ok(self):
        self.result = ("ok", self._df_with_lots)
        self.destroy()

    def _on_no_split(self):
        df = self.df.copy()
        df["ロット"] = 1
        self.result = ("ok", df)
        self.destroy()

    def _on_cancel(self):
        self.result = ("cancel", None)
        self.destroy()

    def _on_lot_double_click(self, event=None):
        selected = self.tree.selection()
        if not selected:
            return
        values = self.tree.item(selected[0], "values")
        if not values:
            return
        try:
            lot_num = int(str(values[1]).replace("ロット", ""))
        except (ValueError, AttributeError):
            return
        if not hasattr(self, "_df_with_lots"):
            return
        lot_df = self._df_with_lots[self._df_with_lots["ロット"] == lot_num].copy()
        if len(lot_df) == 0:
            return
        LotDetailDialog(self, lot_df, lot_num, self.hinshoku_num, self.product_name)


# =========================
# ■ ロット詳細ダイアログ
# =========================
class LotDetailDialog(tk.Toplevel):

    def __init__(self, parent, df_lot, lot_num, hinshoku_num=None, product_name=""):
        super().__init__(parent)
        self.lot_num = lot_num
        self.hinshoku_num = hinshoku_num
        self.product_name = product_name
        self._df_lot = df_lot.copy()

        display = product_name if product_name else (
            f"品種番号{hinshoku_num}" if hinshoku_num is not None else "")
        self.title(f"ロット{lot_num} 詳細  {display}")
        self.geometry("1000x680")
        self.resizable(True, True)

        stats = self._compute_stats()
        if stats is None:
            ttk.Label(
                self,
                text=f"OKデータが {MIN_OK_COUNT} 件未満のため統計計算できません",
                font=("", 12), foreground="#9C0006",
            ).pack(expand=True)
            return

        (self.df_ok, self.mean, self.std, self.ci,
         self.max1, self.min1, self.lower, self.upper,
         self.outliers_df, self.rank_counts,
         self.total_count, self.original_ok_count) = stats

        self._build_ui()

        self.update_idletasks()
        x = parent.winfo_rootx() + 40
        y = parent.winfo_rooty() + 40
        self.geometry(f"+{x}+{y}")

    def _compute_stats(self):
        group = self._df_lot
        rank_counts = group["ランクコード"].value_counts().reset_index()
        rank_counts.columns = ["ランクコード", "件数"]
        rank_counts["内容"] = rank_counts["ランクコード"].map(_RANK_LABEL_MAP)
        rank_counts = rank_counts[["ランクコード", "内容", "件数"]]

        total_count = len(group)
        original_ok_count = int((group["ランクコード"] == "2").sum())

        df_ok = group[group["ランクコード"] == "2"].copy()
        data = pd.to_numeric(df_ok["測定値(g)"], errors="coerce")
        df_ok = df_ok.loc[data.notna()].copy()
        data = data.loc[data.notna()]
        data_arr = np.asarray(data).ravel()

        if len(data_arr) < MIN_OK_COUNT:
            return None

        mean, std, ci, max1, min1, lower, upper = analyze(data_arr)
        outliers_df = df_ok[(df_ok["測定値(g)"] < lower) | (df_ok["測定値(g)"] > upper)]
        return (df_ok, mean, std, ci, max1, min1, lower, upper,
                outliers_df, rank_counts, total_count, original_ok_count)

    def _build_ui(self):
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=10, pady=10)

        tab_sum = ttk.Frame(nb)
        tab_ts  = ttk.Frame(nb)
        tab_ok  = ttk.Frame(nb)
        tab_out = ttk.Frame(nb)
        tab_exp = ttk.Frame(nb)

        nb.add(tab_sum, text="📊 サマリー")
        nb.add(tab_ts,  text="📈 時系列チャート")
        nb.add(tab_ok,  text="✅ OKデータ")
        nb.add(tab_out, text="⚠ 外れ値")
        nb.add(tab_exp, text="💾 エクスポート")

        self._build_summary_tab(tab_sum, Figure, FigureCanvasTkAgg)
        self._build_series_tab(tab_ts, Figure, FigureCanvasTkAgg, NavigationToolbar2Tk)
        self._build_ok_tab(tab_ok)
        self._build_outlier_tab(tab_out)
        self._build_export_tab(tab_exp)

    # ------ タブ1: サマリー ------
    def _build_summary_tab(self, parent, Figure, FigureCanvasTkAgg):
        left = ttk.Frame(parent, padding=10)
        left.pack(side="left", fill="y")

        ttk.Label(left, text="■ 統計値", font=("", 10, "bold")).pack(anchor="w", pady=(0, 5))

        tree = ttk.Treeview(left, columns=("値",), show="tree headings", height=16)
        tree.column("#0", width=160, anchor="w")
        tree.column("値", width=120, anchor="e")
        tree.heading("#0", text="項目")
        tree.heading("値", text="値")

        ng = self.total_count - self.original_ok_count
        dr = (ng / self.total_count * 100) if self.total_count > 0 else 0.0
        rows = [
            ("全数",          f"{self.total_count:,}"),
            ("OK数",          f"{self.original_ok_count:,}"),
            ("NG数",          f"{ng:,}"),
            ("不良率(%)",      f"{dr:.3f}"),
            ("",              ""),
            ("平均(g)",        f"{self.mean:.4f}"),
            ("標準偏差(g)",    f"{self.std:.5f}"),
            ("OKデータ件数",   f"{len(self.df_ok):,}"),
            ("Max(g)",        f"{self.max1:.3f}"),
            ("Min(g)",        f"{self.min1:.3f}"),
            ("下限 −3σ(g)",   f"{self.lower:.4f}"),
            ("上限 +3σ(g)",   f"{self.upper:.4f}"),
            ("",              ""),
            ("95%CI 下限(g)", f"{self.ci[0]:.4f}" if self.ci[0] is not None else "-"),
            ("95%CI 上限(g)", f"{self.ci[1]:.4f}" if self.ci[1] is not None else "-"),
            ("外れ値件数",     f"{len(self.outliers_df):,}"),
        ]
        for label, val in rows:
            tree.insert("", "end", text=label, values=(val,))
        tree.pack(fill="y")

        right = ttk.Frame(parent, padding=10)
        right.pack(side="right", fill="both", expand=True)

        data = pd.to_numeric(self.df_ok["測定値(g)"], errors="coerce").dropna()
        fig = Figure(figsize=(7, 5), dpi=90)
        ax = fig.add_subplot(111)
        ax.hist(data, bins=30, edgecolor="black", alpha=0.7, color="steelblue")
        ax.axvline(self.mean,  color="red",    linestyle="-",  linewidth=2, label=f"平均: {self.mean:.3f}")
        ax.axvline(self.lower, color="orange", linestyle="--", linewidth=2, label=f"-3σ: {self.lower:.3f}")
        ax.axvline(self.upper, color="orange", linestyle="--", linewidth=2, label=f"+3σ: {self.upper:.3f}")
        ax.set_title(f"ロット{self.lot_num} ヒストグラム (n={len(data):,})", fontsize=11, fontweight="bold")
        ax.set_xlabel("測定値(g)")
        ax.set_ylabel("頻度")
        ax.legend(fontsize=9)
        ax.grid(True, alpha=0.3)
        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=right)
        canvas.get_tk_widget().pack(fill="both", expand=True)
        canvas.draw()

    # ------ タブ2: 時系列チャート ------
    def _build_series_tab(self, parent, Figure, FigureCanvasTkAgg, NavigationToolbar2Tk):
        df_ok = self.df_ok
        outlier_mask = (df_ok["測定値(g)"] < self.lower) | (df_ok["測定値(g)"] > self.upper)
        y_vals = df_ok["測定値(g)"].values
        has_datetime = df_ok["日付時刻"].notna().any()

        fig = Figure(figsize=(11, 5), dpi=90)
        ax = fig.add_subplot(111)

        if has_datetime:
            x_all = df_ok["日付時刻"]
            x_ok  = df_ok.loc[~outlier_mask, "日付時刻"]
            x_out = df_ok.loc[outlier_mask,  "日付時刻"]
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
            ax.xaxis.set_major_locator(mdates.AutoDateLocator())
            fig.autofmt_xdate(rotation=30)
            ax.set_xlabel("時刻", fontsize=12)
        else:
            x_all = np.arange(1, len(df_ok) + 1)
            x_ok  = x_all[~outlier_mask.values]
            x_out = x_all[outlier_mask.values]
            ax.set_xlabel("測定順序", fontsize=12)

        y_ok  = y_vals[~outlier_mask.values]
        y_out = y_vals[outlier_mask.values]

        ax.plot(x_all, y_vals, color="steelblue", linewidth=0.6, alpha=0.4, zorder=1)
        ax.scatter(x_ok, y_ok, color="steelblue", s=18, alpha=0.8, zorder=2, label="OK")
        if len(x_out) > 0:
            ax.scatter(x_out, y_out, color="red", s=50, marker="x",
                       linewidths=2, zorder=3, label=f"外れ値 ({len(x_out)}件)")
        ax.axhline(self.mean,  color="red",    linewidth=1.5, linestyle="-",  label=f"平均: {self.mean:.2f}")
        ax.axhline(self.upper, color="orange", linewidth=1.5, linestyle="--", label=f"+3σ: {self.upper:.2f}")
        ax.axhline(self.lower, color="orange", linewidth=1.5, linestyle="--", label=f"-3σ: {self.lower:.2f}")
        ax.set_title(f"ロット{self.lot_num} 時系列チャート (n={len(df_ok):,})", fontsize=12, fontweight="bold")
        ax.set_ylabel("測定値(g)", fontsize=12)
        ax.legend(fontsize=10)
        ax.grid(True, alpha=0.3)
        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=5)
        canvas.draw()

        tb_frame = ttk.Frame(parent)
        tb_frame.pack(fill="x", padx=10)
        NavigationToolbar2Tk(canvas, tb_frame)

    # ------ タブ3: OKデータ ------
    def _build_ok_tab(self, parent):
        outlier_nos = set(self.outliers_df["測定値出力No."].values)
        frame = ttk.Frame(parent, padding=10)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="※ 赤色行が外れ値（±3σ超）です",
                  foreground="#9C0006").pack(anchor="w", pady=(0, 5))

        cols = ("測定値出力No.", "日付時刻", "測定値(g)")
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        tree.tag_configure("outlier", background="#FFCCCC", foreground="#9C0006")
        tree.column("測定値出力No.", anchor="center", width=130)
        tree.column("日付時刻",     anchor="center", width=170)
        tree.column("測定値(g)",    anchor="e",      width=100)
        for col in cols:
            tree.heading(col, text=col)

        for _, row in self.df_ok.iterrows():
            dt  = row["日付時刻"]
            val = row["測定値(g)"]
            dt_str = dt.strftime("%Y-%m-%d %H:%M:%S") if pd.notna(dt) else "-"
            is_out = row["測定値出力No."] in outlier_nos
            tree.insert("", "end",
                        tags=("outlier",) if is_out else (),
                        values=(row["測定値出力No."], dt_str,
                                f"{val:.3f}" if pd.notna(val) else "-"))

        sb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

    # ------ タブ4: 外れ値 ------
    def _build_outlier_tab(self, parent):
        frame = ttk.Frame(parent, padding=10)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text=f"外れ値: {len(self.outliers_df):,}件（±3σ超）",
                  font=("", 10, "bold")).pack(anchor="w", pady=(0, 5))

        cols = ("測定値出力No.", "日付時刻", "測定値(g)")
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        tree.column("測定値出力No.", anchor="center", width=130)
        tree.column("日付時刻",     anchor="center", width=170)
        tree.column("測定値(g)",    anchor="e",      width=100)
        for col in cols:
            tree.heading(col, text=col)

        for _, row in self.outliers_df.iterrows():
            dt  = row["日付時刻"]
            val = row["測定値(g)"]
            dt_str = dt.strftime("%Y-%m-%d %H:%M:%S") if pd.notna(dt) else "-"
            tree.insert("", "end", values=(
                row["測定値出力No."], dt_str,
                f"{val:.3f}" if pd.notna(val) else "-",
            ))

        sb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

    # ------ タブ5: エクスポート ------
    def _build_export_tab(self, parent):
        frame = ttk.Frame(parent, padding=20)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text=f"ロット{self.lot_num} の分析結果をExcelに出力します。",
                  font=("", 11)).pack(anchor="w", pady=(0, 10))
        ttk.Label(frame, text=(
            "出力内容:\n"
            "  ・分析レポート（統計値 + 製造情報 + グラフ）\n"
            "  ・全データ時系列（OK/軽量/過量を色分け）\n"
            "  ・全データ（NG行は赤色）\n"
            "  ・ヒストグラム／時系列チャート（OK品のみ）\n"
            "  ・全OKデータ（外れ値は赤色）\n"
            "  ・全外れ値"
        ), justify="left").pack(anchor="w", pady=(0, 20))

        ttk.Button(frame, text="📤 Excelに出力する",
                   command=self._on_export).pack(anchor="w")

        self._export_status_var = tk.StringVar()
        ttk.Label(frame, textvariable=self._export_status_var,
                  foreground="navy").pack(anchor="w", pady=(10, 0))

    def _on_export(self):
        lot_date = self._df_lot["日付時刻"].dropna().min()
        lot_end  = self._df_lot["日付時刻"].dropna().max()
        if pd.notna(lot_date) and pd.notna(lot_end):
            mfg_start    = lot_date.strftime("%H:%M")
            mfg_end      = lot_end.strftime("%H:%M")
            total_min    = int((lot_end - lot_date).total_seconds() / 60)
            mfg_duration = f"{total_min // 60}時間{total_min % 60:02d}分"
        else:
            mfg_start = mfg_end = mfg_duration = "−"

        display_label = (self.product_name if self.product_name
                         else (f"品種番号{self.hinshoku_num}" if self.hinshoku_num is not None else None))

        if pd.notna(lot_date) and display_label:
            date_str = f"{lot_date.year}/{lot_date.month}/{lot_date.day}"
            chart_prefix = f"{date_str}製造 {display_label} ロット{self.lot_num}　"
            date_str_safe = date_str.replace("/", "-")
            safe_label = display_label.replace("/", "-").replace("\\", "-")
            default_name = (f"分析結果_{date_str_safe}製造_{safe_label} "
                            f"ロット{self.lot_num}_{datetime.datetime.now():%Y%m%d_%H%M}.xlsx")
        else:
            date_str = None
            chart_prefix = f"ロット{self.lot_num}　"
            default_name = f"分析結果_ロット{self.lot_num}_{datetime.datetime.now():%Y%m%d_%H%M}.xlsx"

        filepath = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not filepath:
            return

        try:
            self._export_status_var.set("出力中...")
            self.update_idletasks()

            data = pd.to_numeric(self.df_ok["測定値(g)"], errors="coerce").dropna()

            img_hist = _make_lot_histogram(data, self.mean, self.lower, self.upper, chart_prefix)
            img_series = _make_lot_timeseries(self.df_ok, self.mean, self.lower, self.upper, chart_prefix)
            img_all_series = _make_all_data_timeseries(
                self._df_lot, self.mean, self.std, self.lower, self.upper, chart_prefix)

            save_to_excel(
                self.df_ok, self.mean, self.std, self.ci,
                self.max1, self.min1, self.lower, self.upper,
                self.outliers_df, img_hist, img_series, self.rank_counts,
                filepath, self.lot_num,
                total_count=self.total_count, original_ok_count=self.original_ok_count,
                hinshoku_num=self.hinshoku_num, date_str=date_str,
                product_name=self.product_name,
                df_all=self._df_lot, img_all_series=img_all_series,
                mfg_start=mfg_start, mfg_end=mfg_end, mfg_duration=mfg_duration,
            )
            self._export_status_var.set(f"✓ 出力完了: {os.path.basename(filepath)}")
            messagebox.showinfo("完了", f"Excelファイルを出力しました:\n{filepath}", parent=self)
        except Exception as e:
            import traceback
            self._export_status_var.set("出力失敗")
            messagebox.showerror(
                "エラー", f"出力に失敗しました:\n{e}\n\n{traceback.format_exc()}", parent=self)


# =========================
# ■ 分析
# =========================
def analyze(data):

    data = np.asarray(data).astype(float).ravel()
    data = data[~np.isnan(data)]

    n = len(data)
    if n < 2:
        return None, None, (None, None), None, None, None, None

    mean = float(np.mean(data))
    std = float(np.std(data, ddof=1))

    t_value = stats.t.ppf(0.975, df=n - 1)
    margin = t_value * std / np.sqrt(n)

    max1 = float(np.max(data))
    min1 = float(np.min(data))
    lower = mean - 3 * std
    upper = mean + 3 * std

    return mean, std, (mean - margin, mean + margin), max1, min1, lower, upper


def _make_lot_histogram(data, mean, lower, upper, title_prefix):
    """ヒストグラム画像を生成してBytesIOで返す"""
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.hist(data, bins=30, edgecolor="black", alpha=0.7)
    ax.axvline(mean, color="red", linestyle="-", linewidth=2, label=f"平均: {mean:.3f}")
    ax.axvline(lower, color="orange", linestyle="--", linewidth=2, label=f"下限(-3σ): {lower:.3f}")
    ax.axvline(upper, color="orange", linestyle="--", linewidth=2, label=f"上限(+3σ): {upper:.3f}")
    ax.set_title(f"{title_prefix}測定値の分布（OK品のみ・n={len(data)}）", fontsize=14, fontweight="bold")
    ax.set_xlabel("測定値(g)", fontsize=12)
    ax.set_ylabel("頻度", fontsize=12)
    ax.legend(fontsize=10, loc="upper left", bbox_to_anchor=(1.02, 1), borderaxespad=0)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()

    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=100, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


def _make_lot_timeseries(df_ok, mean, lower, upper, title_prefix):
    """時系列チャート（OK品のみ）画像を生成してBytesIOで返す"""
    fig, ax = plt.subplots(figsize=(12, 5))

    y_vals = df_ok["測定値(g)"].values
    outlier_mask = (df_ok["測定値(g)"] < lower) | (df_ok["測定値(g)"] > upper)
    has_datetime = df_ok["日付時刻"].notna().any()

    if has_datetime:
        x_all = df_ok["日付時刻"]
        x_ok  = df_ok.loc[~outlier_mask, "日付時刻"]
        x_out = df_ok.loc[outlier_mask,  "日付時刻"]
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
        ax.xaxis.set_major_locator(mdates.AutoDateLocator())
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=30, ha="right")
        ax.set_xlabel("時刻", fontsize=12)
    else:
        x_all = np.arange(1, len(df_ok) + 1)
        x_ok  = x_all[~outlier_mask.values]
        x_out = x_all[outlier_mask.values]
        ax.set_xlabel("測定順序", fontsize=12)

    y_ok  = y_vals[~outlier_mask.values]
    y_out = y_vals[outlier_mask.values]

    ax.plot(x_all, y_vals, color="steelblue", linewidth=0.6, alpha=0.4, zorder=1)
    ax.scatter(x_ok, y_ok, color="steelblue", s=18, alpha=0.8, zorder=2, label="OK")
    if len(x_out) > 0:
        ax.scatter(x_out, y_out, color="red", s=50, marker="x",
                   linewidths=2, zorder=3, label=f"外れ値・±3σ超 ({len(x_out)}件)")

        x_out_arr = np.asarray(x_out)
        max_idx = int(np.argmax(y_out))
        min_idx = int(np.argmin(y_out))
        for idx, voffset in [(max_idx, 18), (min_idx, -18)]:
            ax.annotate(f"{y_out[idx]:.3f}", (x_out_arr[idx], y_out[idx]),
                        textcoords="offset points", xytext=(0, voffset),
                        ha="center", va="center", fontsize=9, color="red", fontweight="bold",
                        zorder=6, bbox=dict(boxstyle="round,pad=0.2", facecolor="white",
                                             edgecolor="red", alpha=0.9),
                        arrowprops=dict(arrowstyle="-", color="red", linewidth=0.8))

    ax.axhline(mean,  color="red",    linewidth=1.5, linestyle="-",  label=f"平均: {mean:.3f}")
    ax.axhline(upper, color="orange", linewidth=1.5, linestyle="--", label=f"+3σ: {upper:.3f}")
    ax.axhline(lower, color="orange", linewidth=1.5, linestyle="--", label=f"-3σ: {lower:.3f}")

    ax.set_title(f"{title_prefix}時系列チャート（OK品のみ・n={len(df_ok)}）", fontsize=14, fontweight="bold")
    ax.set_ylabel("測定値(g)", fontsize=12)
    ax.legend(fontsize=10, loc="upper left", bbox_to_anchor=(1.02, 1), borderaxespad=0)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()

    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=100, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


def _make_all_data_timeseries(group, mean, std, lower, upper, title_prefix):
    """全データ時系列チャート（OK/軽量/過量を色分け）画像を生成してBytesIOで返す"""
    group_sorted = group.sort_values("日付時刻").reset_index(drop=True)
    ok_mask    = group_sorted["ランクコード"] == "2"
    kacho_mask = group_sorted["ランクコード"] == "E"
    keiry_mask = group_sorted["ランクコード"] == "1"

    y_vals_all  = pd.to_numeric(group_sorted["測定値(g)"], errors="coerce")
    y_ok_all    = y_vals_all[ok_mask].values
    y_kacho_all = y_vals_all[kacho_mask].values
    y_keiry_all = y_vals_all[keiry_mask].values

    fig, ax = plt.subplots(figsize=(12, 5))

    if group_sorted["日付時刻"].notna().any():
        x_ok_all    = group_sorted.loc[ok_mask,    "日付時刻"]
        x_kacho_all = group_sorted.loc[kacho_mask, "日付時刻"]
        x_keiry_all = group_sorted.loc[keiry_mask, "日付時刻"]
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
        ax.xaxis.set_major_locator(mdates.AutoDateLocator())
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=30, ha="right")
        ax.set_xlabel("時刻", fontsize=12)
    else:
        x_base      = np.arange(1, len(group_sorted) + 1)
        x_ok_all    = x_base[ok_mask.values]
        x_kacho_all = x_base[kacho_mask.values]
        x_keiry_all = x_base[keiry_mask.values]
        ax.set_xlabel("測定順序", fontsize=12)

    ax.plot(x_ok_all, y_ok_all, color="steelblue", linewidth=0.6, alpha=0.4, zorder=1)
    ax.scatter(x_ok_all, y_ok_all, color="steelblue", s=18, alpha=0.8, zorder=2,
               label=f"OK ({ok_mask.sum()}件)")
    if kacho_mask.any():
        ax.scatter(x_kacho_all, y_kacho_all, color="red", s=60, marker="^", zorder=4,
                   label=f"過量 ({kacho_mask.sum()}件)")
    if keiry_mask.any():
        ax.scatter(x_keiry_all, y_keiry_all, color="orange", s=60, marker="v", zorder=4,
                   label=f"軽量 ({keiry_mask.sum()}件)")

    ax.axhline(mean,  color="red",       linewidth=1.5, linestyle="-",  label=f"平均（OK品）: {mean:.3f}")
    ax.axhline(upper, color="darkorange", linewidth=1.5, linestyle="--", label=f"+3σ（OK品）: {upper:.3f}")
    ax.axhline(lower, color="darkorange", linewidth=1.5, linestyle="--", label=f"-3σ（OK品）: {lower:.3f}")

    y_lo = mean - 10 * std
    y_hi = mean + 10 * std
    data_min = y_vals_all.min()
    data_max = y_vals_all.max()
    pad = max((data_max - data_min) * 0.12, std * 2)
    y_lo = min(y_lo, data_min - pad)
    y_hi = max(y_hi, data_max + pad)
    ax.set_ylim(y_lo, y_hi)

    ax.set_title(f"{title_prefix}全データ時系列（n={len(group_sorted)}）", fontsize=14, fontweight="bold")
    ax.set_ylabel("測定値(g)", fontsize=12)
    ax.legend(fontsize=10, loc="upper left", bbox_to_anchor=(1.02, 1), borderaxespad=0)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()

    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=100, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


# =========================
# ■ Excel出力
# =========================
def _create_report_sheet(wb, df_ok, mean, std, ci, max1, min1, lower, upper,
                          outliers_df, img_hist_bytes, img_series_bytes, rank_counts,
                          total_count, original_ok_count, hinshoku_num, date_str, lot,
                          product_name="", img_all_series_bytes=None,
                          mfg_start="−", mfg_end="−", mfg_duration="−"):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.drawing.image import Image
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker

    ws = wb.create_sheet("分析レポート", 0)

    # スタイル
    title_font  = Font(bold=True, size=16, color="FFFFFFFF")
    title_fill  = PatternFill(start_color="FF1F3864", fill_type="solid")
    sec_font    = Font(bold=True, size=10, color="FFFFFFFF")
    sec_fill    = PatternFill(start_color="FF4472C4", fill_type="solid")
    label_font  = Font(bold=True, size=10)
    label_fill  = PatternFill(start_color="FFD9E1F2", fill_type="solid")
    val_fill_e  = PatternFill(start_color="FFFFFFFF", fill_type="solid")
    val_fill_o  = PatternFill(start_color="FFEFF3FB", fill_type="solid")
    warn_fill   = PatternFill(start_color="FFFFF2CC", fill_type="solid")
    warn_font   = Font(bold=True, size=10, color="FF7F6000")
    subheader_fill = PatternFill(start_color="FF8EA9C1", fill_type="solid")
    subheader_font = Font(bold=True, size=9, color="FFFFFFFF")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left", vertical="center", indent=1)

    # 列幅
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 12
    for col in ["E", "F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[col].width = 10.4  # E〜K合計 ≈ 13.5cm

    ng_count    = total_count - original_ok_count
    defect_rate = (ng_count / total_count * 100) if total_count > 0 else 0.0

    # タイトル行
    display_label = product_name if product_name else (f"品種番号 {hinshoku_num}" if hinshoku_num is not None else None)
    if display_label and date_str:
        title = f"{date_str}製造   {display_label}   ロット{lot}   分析レポート"
    else:
        title = f"ロット{lot}   分析レポート"

    ws.merge_cells("A1:J1")
    ws["A1"] = title
    ws["A1"].font      = title_font
    ws["A1"].fill      = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 26

    ws.merge_cells("K1:L1")
    ws["K1"] = datetime.datetime.now().strftime("%Y年%m月%d日")
    ws["K1"].font      = Font(bold=True, size=11, color="FFFFFFFF")
    ws["K1"].fill      = title_fill
    ws["K1"].alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # 統計セクションヘッダー
    ws.merge_cells("A3:B3")
    ws["A3"] = "■ 基本統計"
    ws["A3"].font      = sec_font
    ws["A3"].fill      = sec_fill
    ws["A3"].alignment = center
    ws.row_dimensions[3].height = 16

    # 統計テーブル定義: (ラベル, 値, 書式, 警告フラグ)
    # ラベルが "__subheader__" の行は小見出しとして描画
    stats_rows = [
        ("全数",            total_count,       "0",     False),
        ("OK数",            original_ok_count, "0",     False),
        ("NG数",            ng_count,          "0",     ng_count > 0),
        ("不良率 (%)",       defect_rate,       "0.00",  ng_count > 0),
        ("__subheader__",  "製造情報",          None,    False),
        ("開始時刻",         mfg_start,         "@",     False),
        ("終了時刻",         mfg_end,           "@",     False),
        ("製造時間",         mfg_duration,      "@",     False),
        ("__subheader__",  "統計値（OK品のみ）", None,   False),
        ("平均 (g)",        mean,              "0.000", False),
        ("標準偏差 (g)",     std,               "0.000", False),
        ("Max (g)",        max1,              "0.000", False),
        ("Min (g)",        min1,              "0.000", False),
        ("下限 −3σ (g)",   lower,             "0.000", False),
        ("上限 +3σ (g)",   upper,             "0.000", False),
        (None, None, None, False),
        ("95% CI 下限 (g)",  ci[0] if ci[0] is not None else "−",    "0.000", False),
        ("95% CI 上限 (g)",  ci[1] if ci[1] is not None else "−",    "0.000", False),
        ("外れ値件数",       len(outliers_df),  "0",     len(outliers_df) > 0),
    ]

    data_row = 4
    stripe   = 0
    for label, value, fmt, warn in stats_rows:
        if label is None:
            ws.row_dimensions[data_row].height = 5
            data_row += 1
            continue
        if label == "__subheader__":
            ws.merge_cells(f"A{data_row}:B{data_row}")
            ws[f"A{data_row}"] = value
            ws[f"A{data_row}"].font      = subheader_font
            ws[f"A{data_row}"].fill      = subheader_fill
            ws[f"A{data_row}"].border    = thin
            ws[f"A{data_row}"].alignment = center
            ws.row_dimensions[data_row].height = 13
            data_row += 1
            continue

        ws.row_dimensions[data_row].height = 14
        vfill = warn_fill if warn else (val_fill_e if stripe % 2 == 0 else val_fill_o)
        vfont = warn_font if warn else Font(size=10)

        ws[f"A{data_row}"] = label
        ws[f"A{data_row}"].font      = label_font
        ws[f"A{data_row}"].fill      = label_fill
        ws[f"A{data_row}"].border    = thin
        ws[f"A{data_row}"].alignment = left

        ws[f"B{data_row}"] = value
        ws[f"B{data_row}"].font          = vfont
        ws[f"B{data_row}"].fill          = vfill
        ws[f"B{data_row}"].border        = thin
        ws[f"B{data_row}"].alignment     = center
        ws[f"B{data_row}"].number_format = fmt

        data_row += 1
        stripe   += 1

    ws.row_dimensions[data_row].height = 7  # 統計〜ランク間ギャップ

    # ランクコード集計ミニテーブル
    rank_row = data_row + 1
    ws.merge_cells(f"A{rank_row}:C{rank_row}")
    ws[f"A{rank_row}"] = "■ ランクコード集計"
    ws[f"A{rank_row}"].font      = sec_font
    ws[f"A{rank_row}"].fill      = sec_fill
    ws[f"A{rank_row}"].alignment = center
    ws.row_dimensions[rank_row].height = 16

    rank_header_row = rank_row + 1
    for ci_idx, col_name in enumerate(["内容", "件数", "比率(%)"]):
        cell = ws.cell(row=rank_header_row, column=ci_idx + 1, value=col_name)
        cell.font      = Font(bold=True, color="FFFFFFFF")
        cell.fill      = PatternFill(start_color="FF4472C4", fill_type="solid")
        cell.border    = thin
        cell.alignment = center
    ws.row_dimensions[rank_header_row].height = 14

    for r_idx, row_data in rank_counts[["内容", "件数"]].iterrows():
        r = rank_header_row + 1 + r_idx
        rfill = val_fill_e if r_idx % 2 == 0 else val_fill_o
        ratio = row_data["件数"] / total_count * 100 if total_count > 0 else 0.0
        for c_idx, val in enumerate([row_data["内容"], row_data["件数"], ratio]):
            cell = ws.cell(row=r, column=c_idx + 1, value=val)
            cell.fill          = rfill
            cell.border        = thin
            cell.alignment     = center
            if c_idx == 2:
                cell.number_format = "0.00"
        ws.row_dimensions[r].height = 13

    # ヒストグラム (E3:K27 に TwoCellAnchor で固定 ≈ 13.5cm × 10.5cm)
    img1 = Image(BytesIO(img_hist_bytes))
    img1.width  = 510  # fallback: 13.5cm
    img1.height = 397  # fallback: 10.5cm
    anchor1 = TwoCellAnchor(editAs="twoCell")
    anchor1._from = AnchorMarker(col=4, colOff=0, row=2,  rowOff=0)  # E3  (0-indexed)
    anchor1.to    = AnchorMarker(col=10, colOff=0, row=26, rowOff=0)  # K27 (0-indexed)
    ws.add_image(img1, anchor1)

    # 時系列チャート (下段) ヒストグラムアンカー to=row26(0-indexed) → Excel行27まで占有。行28以降に配置する
    first_series_row = max(data_row, rank_header_row + len(rank_counts) + 1, 26) + 2

    if img_all_series_bytes is not None:
        # ① 全データ時系列
        ws.merge_cells(f"A{first_series_row}:L{first_series_row}")
        ws[f"A{first_series_row}"] = "■ 全データ時系列"
        ws[f"A{first_series_row}"].font      = sec_font
        ws[f"A{first_series_row}"].fill      = sec_fill
        ws[f"A{first_series_row}"].alignment = center
        ws.row_dimensions[first_series_row].height = 16

        img3 = Image(BytesIO(img_all_series_bytes))
        img3.width  = 907
        img3.height = 378
        ws.add_image(img3, f"A{first_series_row + 1}")

        # ② 時系列チャート（OK品のみ）
        series_row = first_series_row + 26
    else:
        series_row = first_series_row

    ws.merge_cells(f"A{series_row}:L{series_row}")
    ws[f"A{series_row}"] = "■ 時系列チャート（OK品のみ）"
    ws[f"A{series_row}"].font      = sec_font
    ws[f"A{series_row}"].fill      = sec_fill
    ws[f"A{series_row}"].alignment = center
    ws.row_dimensions[series_row].height = 16

    img2 = Image(BytesIO(img_series_bytes))
    img2.width  = 907  # 24cm（紙幅いっぱい→fit-to-pageで左右対称に印刷）
    img2.height = 378  # 10cm
    ws.add_image(img2, f"A{series_row + 1}")

    # A4縦・幅1ページ印刷設定（高さは自動）
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.6
    ws.page_margins.right = 0.6
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3


def _style_3col_data_sheet(ws, header_fill, header_font, center_align, border):
    """測定値出力No./日付時刻/測定値(g) の3列データシートに共通スタイルを適用する"""
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 15
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
        for cell in row:
            cell.border = border
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"
    ws.auto_filter.ref = f"A1:C{ws.max_row}"


def save_to_excel(df_ok, mean, std, ci, max1, min1, lower, upper,
                  outliers_df, img_hist, img_series, rank_counts, filename, lot,
                  total_count=0, original_ok_count=0, hinshoku_num=None, date_str=None,
                  product_name="", df_all=None, img_all_series=None,
                  mfg_start="−", mfg_end="−", mfg_duration="−"):

    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.drawing.image import Image

    red_fill    = PatternFill(start_color="FFFF0000", fill_type="solid")
    header_fill = PatternFill(start_color="FF4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    df_all_out = (df_all if df_all is not None else df_ok)[
        ["測定値出力No.", "日付時刻", "測定値(g)", "ランクコード"]
    ].copy()
    df_all_out["ランクコード"] = df_all_out["ランクコード"].map(_RANK_LABEL_MAP).fillna(df_all_out["ランクコード"])

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:

        df_ok[["測定値出力No.", "日付時刻", "測定値(g)"]].to_excel(writer, sheet_name="全OKデータ", index=False)
        outliers_df[["測定値出力No.", "日付時刻", "測定値(g)"]].to_excel(writer, sheet_name="全外れ値（±3σ超）", index=False)
        df_all_out.to_excel(writer, sheet_name="全データ", index=False)

        wb = writer.book

        # ===== OKデータシート =====
        ws_ok = wb["全OKデータ"]
        _style_3col_data_sheet(ws_ok, header_fill, header_font, center_align, border)
        outlier_ids = set(outliers_df["測定値出力No."].values)
        for row in ws_ok.iter_rows(min_row=2, max_row=ws_ok.max_row):
            if row[0].value in outlier_ids:
                for cell in row:
                    cell.fill = red_fill

        # ===== 外れ値シート =====
        ws_out = wb["全外れ値（±3σ超）"]
        _style_3col_data_sheet(ws_out, header_fill, header_font, center_align, border)

        # ===== 全データシート =====
        ws_all = wb["全データ"]
        ws_all.column_dimensions["A"].width = 18
        ws_all.column_dimensions["B"].width = 22
        ws_all.column_dimensions["C"].width = 15
        ws_all.column_dimensions["D"].width = 12
        for cell in ws_all[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        ng_labels = {"軽量", "過量"}
        for row in ws_all.iter_rows(min_row=2, max_row=ws_all.max_row):
            is_ng = row[3].value in ng_labels  # D列 = ランクコード
            for cell in row:
                cell.border = border
                cell.alignment = center_align
                if is_ng:
                    cell.fill = red_fill
            row[1].number_format = "yyyy-mm-dd hh:mm:ss"  # B列 = 日付時刻
        ws_all.auto_filter.ref = f"A1:D{ws_all.max_row}"

        # ===== グラフシート =====
        img_hist_bytes   = img_hist.getvalue()
        img_series_bytes = img_series.getvalue()

        ws_hist = wb.create_sheet("ヒストグラム（OK品のみ）")
        ws_hist.add_image(Image(BytesIO(img_hist_bytes)), "A1")

        ws_series = wb.create_sheet("時系列チャート（OK品のみ）")
        ws_series.add_image(Image(BytesIO(img_series_bytes)), "A1")

        if img_all_series is not None:
            ws_all_series = wb.create_sheet("全データ時系列")
            ws_all_series.add_image(Image(BytesIO(img_all_series.getvalue())), "A1")

        # ===== 分析レポートシート =====
        _create_report_sheet(
            wb, df_ok, mean, std, ci, max1, min1, lower, upper,
            outliers_df, img_hist_bytes, img_series_bytes, rank_counts,
            total_count, original_ok_count, hinshoku_num, date_str, lot,
            product_name=product_name,
            img_all_series_bytes=img_all_series.getvalue() if img_all_series is not None else None,
            mfg_start=mfg_start, mfg_end=mfg_end, mfg_duration=mfg_duration,
        )

        # ===== シート順序を整理 =====
        sheet_order = [
            "分析レポート",
            "全データ時系列",
            "全データ",
            "ヒストグラム（OK品のみ）",
            "時系列チャート（OK品のみ）",
            "全OKデータ",
            "全外れ値（±3σ超）",
        ]
        wb._sheets.sort(
            key=lambda ws: sheet_order.index(ws.title) if ws.title in sheet_order else len(sheet_order)
        )


# =========================
# ■ ロット処理
# =========================
def process_lot(group, lot, save_dir, hinshoku_num=None, product_name=""):
    """
    1ロット分の分析を行いExcelを出力する。
    戻り値:
        ("ok",   ok_count) … 正常に作成
        ("skip", ok_count) … OKデータ不足によりスキップ
    """

    rank_counts = group["ランクコード"].value_counts().reset_index()
    rank_counts.columns = ["ランクコード", "件数"]
    rank_counts["内容"] = rank_counts["ランクコード"].map(_RANK_LABEL_MAP)
    rank_counts = rank_counts[["ランクコード", "内容", "件数"]]

    total_count = len(group)
    original_ok_count = int((group["ランクコード"] == "2").sum())

    df_ok = group[group["ランクコード"] == "2"].copy()

    data = pd.to_numeric(df_ok["測定値(g)"], errors="coerce")
    df_ok = df_ok.loc[data.notna()].copy()
    data = data.loc[data.notna()]
    data = np.asarray(data).ravel()

    if len(data) < MIN_OK_COUNT:
        return ("skip", len(data))

    mean, std, ci, max1, min1, lower, upper = analyze(data)

    outliers_df = df_ok[(df_ok["測定値(g)"] < lower) | (df_ok["測定値(g)"] > upper)]

    lot_date = group["日付時刻"].dropna().min()
    lot_end  = group["日付時刻"].dropna().max()
    if pd.notna(lot_date) and pd.notna(lot_end):
        mfg_start    = lot_date.strftime("%H:%M")
        mfg_end      = lot_end.strftime("%H:%M")
        total_min    = int((lot_end - lot_date).total_seconds() / 60)
        mfg_duration = f"{total_min // 60}時間{total_min % 60:02d}分"
    else:
        mfg_start = mfg_end = mfg_duration = "−"

    display_label = product_name if product_name else (f"品種番号{hinshoku_num}" if hinshoku_num is not None else None)
    if pd.notna(lot_date) and display_label:
        date_str = f"{lot_date.year}/{lot_date.month}/{lot_date.day}"
        chart_prefix = f"{date_str}製造 {display_label} ロット{lot}　"
        date_str_safe = date_str.replace("/", "-")
    else:
        date_str = None
        chart_prefix = f"ロット{lot}　"
        date_str_safe = None

    img_hist = _make_lot_histogram(data, mean, lower, upper, chart_prefix)
    img_series = _make_lot_timeseries(df_ok, mean, lower, upper, chart_prefix)
    img_all_series = _make_all_data_timeseries(group, mean, std, lower, upper, chart_prefix)

    if date_str_safe and display_label:
        safe_label = display_label.replace("/", "-").replace("\\", "-")
        filename = os.path.join(
            save_dir,
            f"分析結果_{date_str_safe}製造_{safe_label} ロット{lot}_{datetime.datetime.now():%Y%m%d_%H%M}.xlsx"
        )
    else:
        filename = os.path.join(
            save_dir,
            f"分析結果_ロット{lot}_{datetime.datetime.now():%Y%m%d_%H%M}.xlsx"
        )

    save_to_excel(df_ok, mean, std, ci, max1, min1,
                  lower, upper, outliers_df,
                  img_hist, img_series, rank_counts, filename, lot,
                  total_count=total_count, original_ok_count=original_ok_count,
                  hinshoku_num=hinshoku_num, date_str=date_str,
                  product_name=product_name,
                  df_all=group, img_all_series=img_all_series,
                  mfg_start=mfg_start, mfg_end=mfg_end, mfg_duration=mfg_duration)

    return ("ok", len(data))


