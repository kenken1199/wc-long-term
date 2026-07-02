"""
==========================================================
 品種別詳細のExcel出力モジュール
==========================================================

 全部入りExcelを生成:
   - 分析レポート（サマリー＋ヒストグラム＋時系列チャート、A4縦1ページ）
   - 統計結果
   - 日別集計
   - 全OKデータ（生データ）
   - 外れ値
==========================================================
"""

from io import BytesIO

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

import font_utils

font_utils.setup_japanese_font()


def _make_histogram(ok_data, mean, std, lower, upper, title_prefix, n):
    """ヒストグラム画像を生成してBytesIOで返す"""
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.hist(ok_data, bins=40, edgecolor="black", alpha=0.7, color="steelblue")
    ax.axvline(mean, color="red", linestyle="-", linewidth=2,
               label=f"平均: {mean:.3f}")
    ax.axvline(lower, color="orange", linestyle="--", linewidth=2,
               label=f"下限(-3σ): {lower:.3f}")
    ax.axvline(upper, color="orange", linestyle="--", linewidth=2,
               label=f"上限(+3σ): {upper:.3f}")
    ax.set_title(f"{title_prefix}測定値の分布(n={n:,})",
                 fontsize=14, fontweight="bold")
    ax.set_xlabel("測定値(g)", fontsize=12)
    ax.set_ylabel("頻度", fontsize=12)
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()

    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=100, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


def _make_timeseries(combined_df, mean, lower, upper, title_prefix):
    """時系列チャート画像を生成してBytesIOで返す"""
    fig, ax = plt.subplots(figsize=(14, 6))

    ok_df = combined_df[combined_df["ランクコード"] == "2"].copy()
    ok_df = ok_df.dropna(subset=["測定値(g)"])

    if len(ok_df) == 0:
        plt.close(fig)
        return None

    y_vals = ok_df["測定値(g)"].values
    outlier_mask = (y_vals < lower) | (y_vals > upper)

    has_dt = ok_df["日付時刻"].notna().any()
    if has_dt:
        x_all = ok_df["日付時刻"]
        x_ok = ok_df.loc[~outlier_mask, "日付時刻"]
        x_out = ok_df.loc[outlier_mask, "日付時刻"]
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d"))
        ax.xaxis.set_major_locator(mdates.AutoDateLocator())
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=30, ha="right")
        ax.set_xlabel("日付", fontsize=12)
    else:
        x_all = np.arange(1, len(ok_df) + 1)
        x_ok = x_all[~outlier_mask]
        x_out = x_all[outlier_mask]
        ax.set_xlabel("測定順序", fontsize=12)

    y_ok = y_vals[~outlier_mask]
    y_out = y_vals[outlier_mask]

    ax.scatter(x_ok, y_ok, color="steelblue", s=8, alpha=0.5, zorder=2, label="OK")
    if len(x_out) > 0:
        ax.scatter(x_out, y_out, color="red", s=30, marker="x",
                   linewidths=1.5, zorder=3, label=f"外れ値({len(x_out)}件)")

    ax.axhline(mean, color="red", linewidth=1.5, linestyle="-",
               label=f"平均: {mean:.3f}")
    ax.axhline(upper, color="orange", linewidth=1.5, linestyle="--",
               label=f"+3σ: {upper:.3f}")
    ax.axhline(lower, color="orange", linewidth=1.5, linestyle="--",
               label=f"-3σ: {lower:.3f}")

    ax.set_title(f"{title_prefix}時系列チャート(n={len(ok_df):,})",
                 fontsize=14, fontweight="bold")
    ax.set_ylabel("測定値(g)", fontsize=12)
    ax.legend(fontsize=10, loc="best")
    ax.grid(True, alpha=0.3)
    fig.tight_layout()

    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=100, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


def _make_daily_trend(daily_df, title_prefix):
    """日別推移グラフ（平均値・σ・不良率の3段）"""
    fig, axes = plt.subplots(3, 1, figsize=(14, 9), sharex=True)

    valid = daily_df.dropna(subset=["平均(g)"])
    if len(valid) == 0:
        plt.close(fig)
        return None

    dates = pd.to_datetime(valid["日付"], format="%Y%m%d", errors="coerce")

    # 平均値
    ax1 = axes[0]
    ax1.plot(dates, valid["平均(g)"], marker="o", color="steelblue", linewidth=1.5)
    overall_mean = valid["平均(g)"].mean()
    ax1.axhline(overall_mean, color="red", linestyle="--", alpha=0.5,
                label=f"全期間平均: {overall_mean:.3f}")
    ax1.set_ylabel("平均(g)", fontsize=11)
    ax1.set_title(f"{title_prefix}日別推移", fontsize=14, fontweight="bold")
    ax1.legend(fontsize=9)
    ax1.grid(True, alpha=0.3)

    # σ
    ax2 = axes[1]
    valid_std = valid.dropna(subset=["σ(g)"])
    dates_std = pd.to_datetime(valid_std["日付"], format="%Y%m%d", errors="coerce")
    ax2.plot(dates_std, valid_std["σ(g)"], marker="s", color="darkorange", linewidth=1.5)
    ax2.set_ylabel("σ(g)", fontsize=11)
    ax2.grid(True, alpha=0.3)

    # 不良率
    ax3 = axes[2]
    ax3.bar(dates, valid["不良率(%)"], color="firebrick", alpha=0.7, width=0.8)
    ax3.set_ylabel("不良率(%)", fontsize=11)
    ax3.set_xlabel("製造日", fontsize=11)
    ax3.grid(True, alpha=0.3)

    for ax in axes:
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y/%m/%d"))
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=30, ha="right")

    fig.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=100, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


def export_hinshoku_detail(filepath, hinshoku_num, aggregate_info,
                            combined_df, daily_df, overall_stats, product_name=""):
    """
    品種別詳細の全部入りExcelを出力する。

    Args:
        filepath: 出力先パス
        hinshoku_num: 品種番号
        aggregate_info: aggregate_by_hinshokuの該当エントリ
        combined_df: 全期間の結合データ（load_hinshoku_dataの結果）
        daily_df: 日別集計（aggregate_by_date_folderの結果）
        overall_stats: compute_overall_statsの結果
        product_name: 製品名（空文字の場合は品種番号で表示）
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.drawing.image import Image

    if overall_stats is None:
        raise ValueError("有効な統計データがありません")

    mean = overall_stats["平均"]
    std = overall_stats["σ"]
    lower = overall_stats["推奨下限"]
    upper = overall_stats["推奨上限"]
    n = overall_stats["件数"]

    display_name = product_name if product_name else f"品種番号{hinshoku_num}"
    title_prefix = f"{display_name} "

    # チャート画像生成
    ok_data = pd.to_numeric(
        combined_df.loc[combined_df["ランクコード"] == "2", "測定値(g)"],
        errors="coerce"
    ).dropna()
    img_hist = _make_histogram(ok_data, mean, std, lower, upper, title_prefix, n)
    img_series = _make_timeseries(combined_df, mean, lower, upper, title_prefix)
    img_trend = _make_daily_trend(daily_df, title_prefix)

    # OKデータと外れ値
    ok_df = combined_df[combined_df["ランクコード"] == "2"].copy()
    ok_df = ok_df.dropna(subset=["測定値(g)"])
    outlier_mask = (ok_df["測定値(g)"] < lower) | (ok_df["測定値(g)"] > upper)
    outliers_df = ok_df[outlier_mask].copy()

    # スタイル準備
    header_fill = PatternFill(start_color="FF4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # ===== 統計結果シート =====
        stats_rows = [
            ("品種番号", hinshoku_num),
        ]
        if product_name:
            stats_rows.append(("製品名", product_name))
        stats_rows += [
            ("初回製造日", aggregate_info.get("初回製造日") or "-"),
            ("最終製造日", aggregate_info.get("最終製造日") or "-"),
            ("製造日数", aggregate_info.get("製造日数", 0)),
            ("ファイル数", aggregate_info.get("ファイル数", 0)),
            ("", ""),
            ("総件数", aggregate_info.get("総件数", 0)),
            ("OK件数", aggregate_info.get("OK件数", 0)),
            ("NG件数", aggregate_info.get("NG件数", 0)),
            ("不良率(%)", round(aggregate_info.get("不良率(%)", 0), 3)),
            ("", ""),
            ("平均(g)", round(mean, 4)),
            ("標準偏差(g)", round(std, 5)),
            ("Min(g)", round(overall_stats["Min"], 3)),
            ("Max(g)", round(overall_stats["Max"], 3)),
            ("推奨下限 -3σ(g)", round(lower, 4)),
            ("推奨上限 +3σ(g)", round(upper, 4)),
            ("95%CI 下限(g)", round(overall_stats["CI下限"], 4)),
            ("95%CI 上限(g)", round(overall_stats["CI上限"], 4)),
            ("外れ値件数", overall_stats["外れ値件数"]),
            ("", ""),
            ("工程実績性能（全体σ使用、≥1.33 推奨）", ""),
            ("Pp",
             round(overall_stats["Pp"], 3) if overall_stats.get("Pp") is not None else "-"),
            ("Ppk",
             round(overall_stats["Ppk"], 3) if overall_stats.get("Ppk") is not None else "-（公称重量未登録）"),
        ]
        stats_df = pd.DataFrame(stats_rows, columns=["項目", "値"])
        stats_df.to_excel(writer, sheet_name="統計結果", index=False)

        # ===== 日別集計シート =====
        daily_export = daily_df.copy()
        for col in ["平均(g)", "σ(g)", "Min(g)", "Max(g)"]:
            if col in daily_export.columns:
                daily_export[col] = daily_export[col].round(4)
        daily_export["不良率(%)"] = daily_export["不良率(%)"].round(3)
        daily_export.to_excel(writer, sheet_name="日別集計", index=False)

        # ===== 全OKデータシート =====
        ok_export = ok_df[["測定値出力No.", "日付時刻", "測定値(g)", "元ファイル"]].copy()
        ok_export.to_excel(writer, sheet_name="全OKデータ", index=False)

        # ===== 外れ値シート =====
        out_export = outliers_df[["測定値出力No.", "日付時刻", "測定値(g)", "元ファイル"]].copy()
        out_export.to_excel(writer, sheet_name="外れ値", index=False)

        wb = writer.book

        # ===== 統計結果シートの整形 =====
        ws = wb["統計結果"]
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 22
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = center
                if cell.value is not None and cell.value != "":
                    cell.border = border

        # ===== 日別集計シートの整形 =====
        ws_d = wb["日別集計"]
        for col_letter in "ABCDEFGHIJ":
            ws_d.column_dimensions[col_letter].width = 13
        for cell in ws_d[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        for row in ws_d.iter_rows(min_row=2, max_row=ws_d.max_row):
            for cell in row:
                cell.alignment = center
                cell.border = border

        # ===== 全OKデータシートの整形 =====
        ws_ok = wb["全OKデータ"]
        ws_ok.column_dimensions["A"].width = 14
        ws_ok.column_dimensions["B"].width = 22
        ws_ok.column_dimensions["C"].width = 13
        ws_ok.column_dimensions["D"].width = 30
        for cell in ws_ok[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        for row in ws_ok.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        if ws_ok.max_row > 1:
            ws_ok.auto_filter.ref = f"A1:D{ws_ok.max_row}"

        # ===== 外れ値シートの整形 =====
        ws_out = wb["外れ値"]
        ws_out.column_dimensions["A"].width = 14
        ws_out.column_dimensions["B"].width = 22
        ws_out.column_dimensions["C"].width = 13
        ws_out.column_dimensions["D"].width = 30
        for cell in ws_out[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        for row in ws_out.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        if ws_out.max_row > 1:
            ws_out.auto_filter.ref = f"A1:D{ws_out.max_row}"

        # ===== グラフシート =====
        ws_h = wb.create_sheet("ヒストグラム")
        ws_h.add_image(Image(BytesIO(img_hist.getvalue())), "A1")

        if img_series:
            ws_s = wb.create_sheet("時系列チャート")
            ws_s.add_image(Image(BytesIO(img_series.getvalue())), "A1")

        if img_trend:
            ws_t = wb.create_sheet("日別推移")
            ws_t.add_image(Image(BytesIO(img_trend.getvalue())), "A1")

        # シート順序: 統計結果, 日別集計, ヒストグラム, 時系列, 日別推移, 全OK, 外れ値
        desired_order = [
            "統計結果", "日別集計", "ヒストグラム", "時系列チャート",
            "日別推移", "全OKデータ", "外れ値"
        ]
        wb._sheets = [wb[name] for name in desired_order if name in wb.sheetnames]
